
# Superclass
from sklearn.metrics import confusion_matrix, classification_report, roc_auc_score
from sklearn.model_selection import train_test_split
from sklearn.utils import class_weight

import reject
from reject.reject import ClassificationRejector
from reject.reject import confusion_matrix as reject_cm
from reject.utils import generate_synthetic_output
from sklearn.metrics import brier_score_loss

import torch
from tqdm import tqdm

from torchvision.transforms import v2
import torchvision
from collections import Counter

import gc

import numpy as np
import pickle
import matplotlib.pyplot as plt
import pandas as pd
import seaborn as sns
import json
import os
import dataframe_image as dfi
import random

import json
import openpyxl
import itertools
import glob
import pathlib
import time
from tempfile import TemporaryDirectory
from copy import deepcopy
import torch.nn as nn

import melanoma as mel


class Model:
    def __init__(self):
        
        pass

    @staticmethod
    def train_model(conf, network, data, dataset_sizes):
        since = time.time()
        # dataset_sizes = {}
        # dataset_sizes['Train'] = data['Train'].__len__()
        # dataset_sizes['Val'] = data['Val'].__len__()

        # Create a temporary directory to save training checkpoints
        with TemporaryDirectory() as tempdir:
            best_model_params_path = os.path.join(tempdir, 'best_model_params.pt')

            torch.save(network.state_dict(), best_model_params_path)
            best_acc = 0.0

            for epoch in range(conf['epochs']):
                print(f"Epoch {epoch}/{conf['epochs'] - 1}")
                print('-' * 10)

                # Each epoch has a training and validation phase
                for phase in ['Train', 'Val']:
                    if phase == 'Train':
                        network.train()  # Set model to training mode
                    else:
                        network.eval()   # Set model to evaluate mode

                    running_loss = 0.0
                    running_corrects = 0

                    # Iterate over data.
                    # for inputs, labels in tqdm(data[phase]):
                    for inputs, labels in tqdm(data[phase], total=len(data[phase])):
                        inputs = inputs.to(conf['device'])
                        # labels = torch.flatten(labels).to(conf['device'])
                        labels = labels.to(conf['device'])

                        # zero the parameter gradients
                        conf['optimizer'].zero_grad()

                        # forward
                        # track history if only in train
                        with torch.set_grad_enabled(phase == 'Train'):
                            outputs = network(inputs)
                            _, preds = torch.max(outputs, 1)
                            loss = conf['criterion'](outputs, labels.type(torch.int64))

                            # backward + optimize only if in training phase
                            if phase == 'Train':
                                loss.backward()
                                conf['optimizer'].step()

                        # statistics
                        running_loss += loss.item() * inputs.size(0)
                        running_corrects += torch.sum(preds == labels.data)
                    if phase == 'Train':
                        conf['scheduler'].step()

                    epoch_loss = running_loss / dataset_sizes[phase]
                    epoch_acc = running_corrects.double() / dataset_sizes[phase]

                    print(f'{phase} Loss: {epoch_loss:.4f} Acc: {epoch_acc:.4f}')

                    # deep copy the model
                    if phase == 'Val' and epoch_acc > best_acc:
                        best_acc = epoch_acc
                        torch.save(network.state_dict(), best_model_params_path)

                print()

            time_elapsed = time.time() - since
            print(f'Training complete in {time_elapsed // 60:.0f}m {time_elapsed % 60:.0f}s')
            print(f'Best val Acc: {best_acc:4f}')

            # load best model weights
            network.load_state_dict(torch.load(best_model_params_path))
            isExist = os.path.exists(conf['snapshot_path'])
            if not isExist :
                os.makedirs(conf['snapshot_path'])
            else:
                pass
            torch.save(network.state_dict(), os.path.join(conf['snapshot_path'], f"{conf['model_file_name']}.pt"))
        return network

    @staticmethod
    def evaluate_model(model, dataloader, device):
        all_labels = {
            'Val': [],
            'Test': [],
        }
        all_preds = {
            'Val': [],
            'Test': [],
        }
        all_scores = {
            'Val': [],
            'Test': [],
        }
        all_paths = {
            'Val': [],
            'Test': [],
        }
        all_ids = {
            'Val': [],
            'Test': [],
        }
        CM=0
        model.to(device)
        # model.eval()
        # num_iters = len(dataloaders['Test'].dataset) / dataloaders['Test'].batch_size
        with torch.no_grad():
            for phase in ['Val', 'Test']:
                for data in tqdm(dataloader[phase]):
                    

                    images, labels, paths = data
                    images = images.to(device)
                    # labels = labels.to(device)
                    
                    outputs = model(images) #file_name
                    smx = nn.Softmax(dim=1)
                    smx_output = smx(outputs.data)
                    # preds = torch.argmax(outputs.data, 1)
                    preds = torch.argmax(smx_output, 1)
                    CM+=confusion_matrix(labels, preds.cpu(),labels=[0,1])

                    all_preds[phase].extend(preds.cpu().tolist())
                    all_labels[phase].extend(labels.tolist())
                    all_scores[phase].extend(smx_output.cpu().tolist())
                    all_paths[phase].extend(paths)

                    assert len(all_preds[phase]) == len(all_labels[phase]) and \
                        len(all_labels[phase]) == len(all_scores[phase]) and \
                        len(all_scores[phase]) == len(all_paths[phase])
                    
                
                    all_ids[phase] = [str(pathlib.Path(path).stem) for path in all_paths[phase]]
                if (phase == 'Test'):
                    tn=CM[0][0]
                    tp=CM[1][1]
                    fp=CM[0][1]
                    fn=CM[1][0]
                    acc=np.sum(np.diag(CM)/np.sum(CM))
                    sensitivity=tp/(tp+fn)
                    precision=tp/(tp+fp)
                    
                    print('\nTestset Accuracy(mean): %f %%' % (100 * acc))
                    print()
                    print('Confusion Matirx : ')
                    print(CM)
                    print('- Sensitivity : ',(tp/(tp+fn))*100)
                    print('- Specificity : ',(tn/(tn+fp))*100)
                    print('- Precision: ',(tp/(tp+fp))*100)
                    print('- NPV: ',(tn/(tn+fn))*100)
                    print('- F1 : ',((2*sensitivity*precision)/(sensitivity+precision))*100)
                    print()

        # Val set
        prob_pos_val = [x[1] for x in all_scores['Val']]
            
        b_score_val = brier_score_loss(all_labels['Val'], prob_pos_val)
        ece_val = mel.Model.expected_calibration_error(all_scores['Val'], all_labels['Val'], M=10)

        # Test set
        test_report = classification_report(all_labels['Test'], all_preds['Test'], target_names=mel.Parser.common_binary_label.values(), output_dict = True)
        mal_prob = [x[1] for x in all_scores['Test']]

        b_score_test = brier_score_loss(all_labels['Test'], mal_prob)
        ece_test = mel.Model.expected_calibration_error(all_scores['Test'], all_labels['Test'], M=10)

        performance = {
            'Val': {
                'y_labels': all_labels['Val'],
                'y_scores': all_scores['Val'],
                'brier_score': b_score_val,
                'ece': ece_val,
            },
            'Test': {
                'y_labels': all_labels['Test'],
                'y_pred': all_preds['Test'],
                'y_scores': all_scores['Test'],
                'y_ids': all_ids['Test'],
                'accuracy': test_report['accuracy'],
                'precision': test_report['macro avg']['precision'],
                'sensitivity': test_report['malignant']['recall'],
                'specificity': test_report['benign']['recall'],
                'f1-score': test_report['macro avg']['f1-score'],
                'auc-roc': roc_auc_score(all_labels['Test'], mal_prob),
                'brier_score': b_score_test,
                'ece': ece_test,
            }
        }
            
                    
        return performance
    
    @staticmethod
    def evaluate_leaderboard(model, model_path, db_path, device):
        all_preds = []
        all_scores = []
        all_ids = []

        data_transform = {
                'Test': v2.Compose([
                v2.Resize(256),
                v2.CenterCrop(224),
                v2.ToTensor(),
                v2.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
            ]),
        }

        # name of csv file
        csv_path = os.path.join(pathlib.Path(model_path).parent, 'leaderboard')
        full_filename = pathlib.Path(model_path).stem

        if not os.path.exists(csv_path):
            os.makedirs(csv_path)
        csv_filename = f'{full_filename}_leaderboard.csv'


        if os.path.exists(os.path.join(csv_path, csv_filename)) is False:

            paths = {
                'ISIC2020': os.path.join(db_path, '..', mel.DatasetType.ISIC2020.name, 'ISIC_2020_Test_Input'),
            }

            datafolders = {
                'ISIC2020': {
                    'Test': mel.ImageDataset(root_dir=paths['ISIC2020'], transform=data_transform['Test'])
                },
            }

            dataloader = {
                'ISIC2020': {
                    'Test': torch.utils.data.DataLoader(datafolders['ISIC2020']['Test'], batch_size=32,
                                                                shuffle=False, pin_memory=True,
                                                                num_workers=4, prefetch_factor=2)
                },
                
            }

            model.to(device)
        
            with torch.no_grad():
                for data in tqdm(dataloader['ISIC2020']['Test']):
                    

                    images, filenames = data
                    images = images.to(device)
                    
                    
                    outputs = model(images) #file_name
                    smx = nn.Softmax(dim=1)
                    smx_output = smx(outputs.data)
                    # preds = torch.argmax(outputs.data, 1)
                    preds = torch.argmax(smx_output, 1)
                    
                    all_ids.extend(filenames)
                    all_preds.extend(preds.cpu().tolist())
                    all_scores.extend(smx_output.cpu().tolist())

                    assert len(all_preds) == len(all_ids) and len(all_preds) == len(all_scores) 


            import csv
            # field names
            fields = ['image_name', 'target']
            
        
        
        
            with open(os.path.join(csv_path, csv_filename), 'w') as csvfile:
                # creating a csv writer object
                csvwriter = csv.writer(csvfile)
            
                # writing the fields
                csvwriter.writerow(fields)

                # writing the data rows
                assert len(all_ids) == len(all_scores)
                for idx, label in enumerate(all_scores):
                    csvwriter.writerow([all_ids[idx], all_scores[idx][1]])
        else:
            print(f'Skipping {csv_filename}')
    @staticmethod
    def evaluate_model_onAll(model, model_path, db_path, device):
        data_transform = {
                'Test': v2.Compose([
                v2.Resize(256),
                v2.CenterCrop(224),
                v2.ToTensor(),
                v2.Normalize([0.485, 0.456, 0.406], [0.229, 0.224, 0.225])
            ]),
        }

        paths = {
            'Val': {
                # 'HAM10000': os.path.join(db_path, mel.DatasetType.HAM10000.name, 'final', 'Test'),
                # 'ISIC2016': os.path.join(db_path, mel.DatasetType.ISIC2016.name, 'final', 'Test'),
                'ISIC2017': os.path.join(db_path, mel.DatasetType.ISIC2017.name, 'final', 'Val'),
                'ISIC2018': os.path.join(db_path, mel.DatasetType.ISIC2018.name, 'final', 'Val'),
                'KaggleMB': os.path.join(db_path, mel.DatasetType.KaggleMB.name, 'final', 'Val'),
                '_7_point_criteria': os.path.join(db_path, mel.DatasetType._7_point_criteria.name, 'final', 'Val')
            },
            'Test': {
                # 'HAM10000': os.path.join(db_path, mel.DatasetType.HAM10000.name, 'final', 'Test'),
                # 'ISIC2016': os.path.join(db_path, mel.DatasetType.ISIC2016.name, 'final', 'Test'),
                'ISIC2017': os.path.join(db_path, mel.DatasetType.ISIC2017.name, 'final', 'Test'),
                'ISIC2018': os.path.join(db_path, mel.DatasetType.ISIC2018.name, 'final', 'Test'),
                'KaggleMB': os.path.join(db_path, mel.DatasetType.KaggleMB.name, 'final', 'Test'),
                '_7_point_criteria': os.path.join(db_path, mel.DatasetType._7_point_criteria.name, 'final', 'Test')
            },

        }

        datafolders = {
            # 'HAM10000': {
            #     'Val': mel.ImageFolder_filename(paths['Val']['HAM10000'], data_transform['Test']),
            #     'Test': mel.ImageFolder_filename(paths['Test']['HAM10000'], data_transform['Test'])
            # },
            # 'ISIC2016': {
            #     'Test': mel.ImageFolder_filename(paths['Test']['ISIC2016'], data_transform['Test'])
            # },
            'ISIC2017': {
                'Val': mel.ImageFolder_filename(paths['Val']['ISIC2017'], data_transform['Test']),
                'Test': mel.ImageFolder_filename(paths['Test']['ISIC2017'], data_transform['Test'])
            },
            'ISIC2018': {
                'Val': mel.ImageFolder_filename(paths['Val']['ISIC2018'], data_transform['Test']),
                'Test': mel.ImageFolder_filename(paths['Test']['ISIC2018'], data_transform['Test'])
            },
            '_7_point_criteria': {
                'Val': mel.ImageFolder_filename(paths['Val']['_7_point_criteria'], data_transform['Test']),
                'Test': mel.ImageFolder_filename(paths['Test']['_7_point_criteria'], data_transform['Test'])
            },
            'KaggleMB': {
                'Val': mel.ImageFolder_filename(paths['Val']['KaggleMB'], data_transform['Test']),
                'Test': mel.ImageFolder_filename(paths['Test']['KaggleMB'], data_transform['Test'])
            }
        }

        dataloaders = {
            # 'HAM10000': {
            #     'Test': torch.utils.data.DataLoader(datafolders['HAM10000']['Test'], batch_size=32,
            #                                                 shuffle=False, pin_memory=True
            #                                                 ,num_workers=4, prefetch_factor=2)
            # },
            # 'ISIC2016': {
            #     'Test': torch.utils.data.DataLoader(datafolders['ISIC2016']['Test'], batch_size=32,
            #                                                 shuffle=False, pin_memory=True,
            #                                                 num_workers=4, prefetch_factor=2)
            # },
            'ISIC2017': {
                'Val': torch.utils.data.DataLoader(datafolders['ISIC2017']['Val'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2),
                'Test': torch.utils.data.DataLoader(datafolders['ISIC2017']['Test'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2)
            },
            'ISIC2018': {
                'Val': torch.utils.data.DataLoader(datafolders['ISIC2018']['Val'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2),
                'Test': torch.utils.data.DataLoader(datafolders['ISIC2018']['Test'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2)
            },
            '_7_point_criteria': {
                'Val': torch.utils.data.DataLoader(datafolders['_7_point_criteria']['Val'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2),
                'Test': torch.utils.data.DataLoader(datafolders['_7_point_criteria']['Test'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2)
            },
            'KaggleMB': {
                'Val': torch.utils.data.DataLoader(datafolders['KaggleMB']['Val'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2),
                'Test': torch.utils.data.DataLoader(datafolders['KaggleMB']['Test'], batch_size=32,
                                                            shuffle=False, pin_memory=True,
                                                            num_workers=4, prefetch_factor=2)
            },
            
        }

        performances = {
            # 'HAM10000': None,
            # 'ISIC2016': None,
            'ISIC2017': None,
            'ISIC2018': None,
            '_7_point_criteria': None,
            'KaggleMB': None,
        }
        json_path = os.path.join(pathlib.Path(model_path).parent, 'performance')
        full_filename = pathlib.Path(model_path).stem
        json_filename = f'{full_filename}_metrics.json'

        if os.path.exists(os.path.join(json_path, json_filename)) is False:
            
            if not os.path.exists(json_path):
                os.makedirs(json_path, exist_ok=True)
            

            classifier_name = pathlib.Path(model_path).parent.name
            

            for db in performances:
                print(f'Evaluating {full_filename} model on {mel.DatasetType[db].name}...\n')
                performances[db] = mel.Model.evaluate_model(model, dataloaders[db], device)    

            DBtypes = [db.name for db in mel.DatasetType]

            def calculate_filesize(model):
                param_size = 0
                for param in model.parameters():
                    param_size += param.nelement() * param.element_size()
                buffer_size = 0
                for buffer in model.buffers():
                    buffer_size += buffer.nelement() * buffer.element_size()

                size_all_mb = (param_size + buffer_size) / 1024**2
                return size_all_mb
            

            used_DBs = [each_model for each_model in DBtypes if(each_model in full_filename)]
            final_perf = {
                'dataset': used_DBs,
                'classifier': classifier_name,
                'Parameters': sum(p.numel() for p in model.parameters()),
                # 'HAM10000': performances['HAM10000'],
                # 'ISIC2016': performances['ISIC2016'],
                'ISIC2017': performances['ISIC2017'],
                'ISIC2018': performances['ISIC2018'],
                'KaggleMB': performances['KaggleMB'],
                '_7_point_criteria': performances['_7_point_criteria'],
                'Filesize': '{:.2f}'.format(calculate_filesize(model)),

            }
            # Dump Json
            file_json = open(os.path.join(json_path, json_filename), "w")

            json.dump(final_perf, file_json, indent = 6)
            
            file_json.close()

            
        else:
            print(f'Skipping {full_filename}')


        


    @staticmethod
    def extract_performances(snapshot_path):
        jsonfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/*/performance/*_*_metrics.json', recursive=True)]))
        jsonnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, jsonfiles))

        final_perf = []

        for idx, j in enumerate(jsonfiles):
            fi = open(j)
            jfile = json.load(fi)
            final_perf.append(jfile)


        performance = openpyxl.Workbook()
        performance_ws = performance.active
        performance_ws.title = 'ISIC2017'
        # performance.create_sheet('ISIC2016')
        # performance.create_sheet('ISIC2017')
        performance.create_sheet('ISIC2018')
        performance.create_sheet('KaggleMB')
        performance.create_sheet('7pointcriteria')


        cols = ['Network', 'DB Comb', 'Precision', 'Specificity', 'Sensitivity', 'F-1 score', 'Accuracy', 'AUC-ROC', 'Brier score', 'ECE', 'Filesize', 'Parameters']

        # HAM10000_ws = performance['HAM10000']
        # HAM10000_ws.append(cols)

        # for p in final_perf:
        #     HAM10000_ws.append([p['classifier'], str(p['dataset']), p['HAM10000']['Test']['precision'], p['HAM10000']['Test']['specificity'], p['HAM10000']['Test']['sensitivity'], p['HAM10000']['Test']['f1-score'], p['HAM10000']['Test']['accuracy'], p['HAM10000']['Test']['auc-roc'], p['Filesize'], p['Parameters']])
            
        # ISIC2016_ws = performance['ISIC2016']
        # ISIC2016_ws.append(cols)

        # for p in final_perf:
        #     ISIC2016_ws.append([p['classifier'], str(p['dataset']), p['ISIC2016']['precision'], p['ISIC2016']['specificity'], p['ISIC2016']['sensitivity'], p['ISIC2016']['f1-score'],  p['ISIC2016']['accuracy'], p['ISIC2016']['auc-roc'], p['Filesize'], p['Parameters']])

        ISIC2017_ws = performance['ISIC2017']
        ISIC2017_ws.append(cols)

        for p in final_perf:
            ISIC2017_ws.append([p['classifier'], str(p['dataset']), p['ISIC2017']['Test']['precision'], 
            p['ISIC2017']['Test']['specificity'], p['ISIC2017']['Test']['sensitivity'], p['ISIC2017']['Test']['f1-score'], 
            p['ISIC2017']['Test']['accuracy'], p['ISIC2017']['Test']['auc-roc'], p['ISIC2017']['Test']['brier_score'], p['ISIC2017']['Test']['ece'],
            p['Filesize'], p['Parameters']])

        ISIC2018_ws = performance['ISIC2018']
        ISIC2018_ws.append(cols)

        for p in final_perf:
            ISIC2018_ws.append([p['classifier'], str(p['dataset']), p['ISIC2018']['Test']['precision'], 
            p['ISIC2018']['Test']['specificity'], p['ISIC2018']['Test']['sensitivity'], p['ISIC2018']['Test']['f1-score'], 
            p['ISIC2018']['Test']['accuracy'], p['ISIC2018']['Test']['auc-roc'],  p['ISIC2018']['Test']['brier_score'], p['ISIC2018']['Test']['ece'],
            p['Filesize'], p['Parameters']])


        KaggleMB_ws = performance['KaggleMB']
        KaggleMB_ws.append(cols)

        for p in final_perf:
            KaggleMB_ws.append([p['classifier'], str(p['dataset']), p['KaggleMB']['Test']['precision'], 
            p['KaggleMB']['Test']['specificity'], p['KaggleMB']['Test']['sensitivity'], p['KaggleMB']['Test']['f1-score'], 
            p['KaggleMB']['Test']['accuracy'], p['KaggleMB']['Test']['auc-roc'],  p['KaggleMB']['Test']['brier_score'], p['KaggleMB']['Test']['ece'],
            p['Filesize'], p['Parameters']])

        _7pointcriteria_ws = performance['7pointcriteria']
        _7pointcriteria_ws.append(cols)

        for p in final_perf:
            _7pointcriteria_ws.append([p['classifier'], str(p['dataset']), p['_7_point_criteria']['Test']['precision'], 
            p['_7_point_criteria']['Test']['specificity'], p['_7_point_criteria']['Test']['sensitivity'], p['_7_point_criteria']['Test']['f1-score'], 
            p['_7_point_criteria']['Test']['accuracy'], p['_7_point_criteria']['Test']['auc-roc'],  p['_7_point_criteria']['Test']['brier_score'], p['_7_point_criteria']['Test']['ece'],
            p['Filesize'], p['Parameters']])
            
        performance.save(f'{snapshot_path}/performance_pytorch.xlsx')

        print(f'{snapshot_path}/performance_pytorch.xlsx' + ' generated')

    @staticmethod
    def extract_positives_per_sample(snapshot_path):
        jsonfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/*/performance/*_metrics.json', recursive=True)]))
        jsonnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, jsonfiles))


        
        test_sets = ['ISIC2017', 'ISIC2018', '_7_point_criteria', 'KaggleMB']
        # test_sets = ['ISIC2017', 'ISIC2018', '_7_point_criteria']
        samples = {}
        
            

        for idx, j in tqdm(enumerate(jsonfiles)):
            fi = open(j)
            jfile = json.load(fi)
            
            
            for db in test_sets:
                samples.setdefault(db, {})
                
                samples[db].setdefault('y_ids', {})
                samples[db].setdefault('y_pred', jfile[db]['Test']['y_pred'])

                samples[db].setdefault('total_miss', [])

                assert len(jfile[db]['Test']['y_pred']) == len(jfile[db]['Test']['y_ids'])
                assert len(jfile[db]['Test']['y_pred']) == len(jfile[db]['Test']['y_labels'])
                
                for i, id in enumerate(jfile[db]['Test']['y_ids']):
                    # Kaggle DB has duplicate filenames in benign/malignant folders, so id_label is created
                    id_label = f"{id}_gt{jfile[db]['Test']['y_labels'][i]}"
                    samples[db]['y_ids'].setdefault(id_label, {})
                    # samples[db]['y_ids'][id].setdefault('y_pred_cor_cnt', 0)
                    # samples[db]['y_ids'][id].setdefault('y_pred_wrg_cnt', 0)
                    samples[db]['y_ids'][id_label].setdefault('classifiers_cor', {})
                    samples[db]['y_ids'][id_label].setdefault('classifiers_wrg', {})
                    # samples[db]['y_ids'].setdefault('y_pred', {})
                    samples[db]['y_ids'][id_label]['y_label'] = jfile[db]['Test']['y_labels'][i]
                    samples[db]['y_ids'][id_label]['y_pred'] = jfile[db]['Test']['y_pred'][i]

                    samples[db]['y_ids'][id_label]['classifiers_cor'].setdefault('count', 0)
                    samples[db]['y_ids'][id_label]['classifiers_wrg'].setdefault('count', 0)
                    samples[db]['y_ids'][id_label]['classifiers_cor'].setdefault('classifiers', [])
                    samples[db]['y_ids'][id_label]['classifiers_wrg'].setdefault('classifiers', [])
                    samples[db]['y_ids'][id_label]['classifiers_cor'].setdefault('statistics', {})
                    samples[db]['y_ids'][id_label]['classifiers_wrg'].setdefault('statistics', {})
                    

                    # if jfile[db]['Test']['y_pred'][i] == jfile[db]['Test']['y_labels'][i]:
                    #     samples[db]['y_ids'][id]['classifiers_cor']['count'] += 1
                    #     samples[db]['y_ids'][id]['classifiers_cor'].append(jfile['classifier'])
                    # elif jfile[db]['Test']['y_pred'][i] != jfile[db]['Test']['y_labels'][i]:
                    #     samples[db]['y_ids'][id]['classifiers_wrg']['count'] += 1
                    #     samples[db]['y_ids'][id]['classifiers_wrg'].append(jfile['classifier'])
                    #     if(samples[db]['y_ids'][id]['classifiers_wrg']['count'] == len(jsonfiles)):
                    #         samples[db]['total_miss'].append(id)
                    if samples[db]['y_ids'][id_label]['y_label'] == samples[db]['y_ids'][id_label]['y_pred']:
                        samples[db]['y_ids'][id_label]['classifiers_cor']['count'] += 1
                        samples[db]['y_ids'][id_label]['classifiers_cor']['classifiers'].append(jfile['classifier'])
                    elif samples[db]['y_ids'][id_label]['y_label'] != samples[db]['y_ids'][id_label]['y_pred']:
                        samples[db]['y_ids'][id_label]['classifiers_wrg']['count'] += 1
                        samples[db]['y_ids'][id_label]['classifiers_wrg']['classifiers'].append(jfile['classifier'])
                        if(samples[db]['y_ids'][id_label]['classifiers_wrg']['count'] == len(jsonfiles)):
                            samples[db]['total_miss'].append(id_label)
                    else:
                        raise AssertionError("Unexpected error")

            # samples[db]['y_ids'][id]['classifiers_cor']['statistics'] = dict(Counter(samples[db]['y_ids'][id]['classifiers_cor']['classifiers']))
            # samples[db]['y_ids'][id]['classifiers_wrg']['statistics'] = dict(Counter(samples[db]['y_ids'][id]['classifiers_wrg']['classifiers']))

        for db in test_sets:
            for id in samples[db]['y_ids']:
                assert len(jsonfiles) == \
                    (samples[db]['y_ids'][id]['classifiers_cor']['count'] + samples[db]['y_ids'][id]['classifiers_wrg']['count'])
                assert mel.CommonData().dbNumImgs[mel.DatasetType[db]]['testimages'] == len(samples[db]['y_ids'])

                samples[db]['y_ids'][id]['classifiers_cor']['statistics'] = dict(Counter(samples[db]['y_ids'][id]['classifiers_cor']['classifiers']))
                samples[db]['y_ids'][id]['classifiers_wrg']['statistics'] = dict(Counter(samples[db]['y_ids'][id]['classifiers_wrg']['classifiers']))



        s = openpyxl.Workbook()
        samples_ws = s.active
        samples_ws.title = 'ISIC2017'
        s.create_sheet('ISIC2018')
        s.create_sheet('KaggleMB')
        s.create_sheet('7pointcriteria')


        cols = ['filename', 'actual label', 'pred_cor_cnt', 'pred_wrg_cnt', 'classifiers_cor', 'classifiers_wrg']

        ws = {}
        for t in test_sets:
            if t == '_7_point_criteria':
                ws[t] = s['7pointcriteria']
            else:
                ws[t] = s[t]
            ws[t].append(cols)
            for id in samples[t]['y_ids']:
                ws[t].append([id, samples[t]['y_ids'][id]['y_label'], samples[t]['y_ids'][id]['classifiers_cor']['count'], samples[t]['y_ids'][id]['classifiers_wrg']['count'], str(samples[t]['y_ids'][id]['classifiers_cor']['statistics']), str(samples[t]['y_ids'][id]['classifiers_wrg']['statistics'])])

            
        s.save(f'{snapshot_path}/benchmarks_analysis.xlsx')

        print(f'{snapshot_path}/benchmarks_analysis.xlsx' + ' generated')


        print('stop')

        return samples[t]['y_ids'][id]['classifiers_cor']['statistics']

    @staticmethod
    def extract_reject_performances(snapshot_path):
        jsonfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/*/performance/*_metrics_reject.json', recursive=True)]))
        jsonnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, jsonfiles))

        final_perf = []

        for idx, j in enumerate(jsonfiles):
            fi = open(j)
            jfile = json.load(fi)
            final_perf.append(jfile)


        performance = openpyxl.Workbook()
        performance_ws = performance.active
        performance_ws.title = 'ISIC2017'
        # performance.create_sheet('ISIC2016')
        # performance.create_sheet('ISIC2017')
        performance.create_sheet('ISIC2018')
        performance.create_sheet('KaggleMB')
        performance.create_sheet('7pointcriteria')


        cols = ['Network', 'DB Comb', 'Precision', 'Specificity', 'Sensitivity', 'F-1 score', 'Accuracy', 'AUC-ROC', 'Threshold', 'Brier score', 'ECE']

        # HAM10000_ws = performance['HAM10000']
        # HAM10000_ws.append(cols)

        

        # for p in final_perf:
        #     HAM10000_ws.append([p['classifier'], str(p['dataset']), p['HAM10000']['Non-rejected']['precision'], 
        #     p['HAM10000']['Non-rejected']['specificity'], 
        #     p['HAM10000']['Non-rejected']['sensitivity'] if isinstance(p['HAM10000']['Non-rejected']['sensitivity'], (int, float)) else 'N/A', 
        #     p['HAM10000']['Non-rejected']['f1-score'], 
        #     p['HAM10000']['Non-rejected']['accuracy'],
        #     p['HAM10000']['Non-rejected']['auc-roc']])
            
        # ISIC2016_ws = performance['ISIC2016']
        # ISIC2016_ws.append(cols)

        # for p in final_perf:
        #     ISIC2016_ws.append([p['classifier'], str(p['dataset']), p['ISIC2016']['Non-rejected']['precision'],
        #     p['ISIC2016']['Non-rejected']['specificity'], 
        #     p['ISIC2016']['Non-rejected']['sensitivity'] if isinstance(p['ISIC2016']['Non-rejected']['sensitivity'], (int, float)) else 'N/A', 
        #     p['ISIC2016']['Non-rejected']['f1-score'], 
        #     p['ISIC2016']['Non-rejected']['accuracy'],
        #     p['ISIC2016']['Non-rejected']['auc-roc']])

        ISIC2017_ws = performance['ISIC2017']
        ISIC2017_ws.append(cols)

        for p in final_perf:
            ISIC2017_ws.append([p['classifier'], str(p['dataset']), p['ISIC2017']['Non-rejected']['precision'], 
            p['ISIC2017']['Non-rejected']['specificity'], 
            p['ISIC2017']['Non-rejected']['sensitivity'] if isinstance(p['ISIC2017']['Non-rejected']['sensitivity'], (int, float)) else 'N/A', 

            p['ISIC2017']['Non-rejected']['f1-score'], 
            p['ISIC2017']['Non-rejected']['accuracy'],
            p['ISIC2017']['Non-rejected']['auc-roc'],
            p['ISIC2017']['Threshold'],
            p['ISIC2017']['brier_score'],
            p['ISIC2017']['ece']])

        ISIC2018_ws = performance['ISIC2018']
        ISIC2018_ws.append(cols)

        for p in final_perf:
            ISIC2018_ws.append([p['classifier'], str(p['dataset']), p['ISIC2018']['Non-rejected']['precision'], 
            p['ISIC2018']['Non-rejected']['specificity'], 
            p['ISIC2018']['Non-rejected']['sensitivity'] if isinstance(p['ISIC2018']['Non-rejected']['sensitivity'], (int, float)) else 'N/A', 
            p['ISIC2018']['Non-rejected']['f1-score'], 
            p['ISIC2018']['Non-rejected']['accuracy'],
            p['ISIC2018']['Non-rejected']['auc-roc'],
            p['ISIC2018']['Threshold'],
            p['ISIC2018']['brier_score'],
            p['ISIC2018']['ece']])


        KaggleMB_ws = performance['KaggleMB']
        KaggleMB_ws.append(cols)

        for p in final_perf:
            KaggleMB_ws.append([p['classifier'], str(p['dataset']), p['KaggleMB']['Non-rejected']['precision'], 
            p['KaggleMB']['Non-rejected']['specificity'], 
            p['KaggleMB']['Non-rejected']['sensitivity'] if isinstance(p['KaggleMB']['Non-rejected']['sensitivity'], (int, float)) else 'N/A', 
            p['KaggleMB']['Non-rejected']['f1-score'], 
            p['KaggleMB']['Non-rejected']['accuracy'],
            p['KaggleMB']['Non-rejected']['auc-roc'],
            p['KaggleMB']['Threshold'],
            p['KaggleMB']['brier_score'],
            p['KaggleMB']['ece']])

        _7pointcriteria_ws = performance['7pointcriteria']
        _7pointcriteria_ws.append(cols)

        for p in final_perf:
            _7pointcriteria_ws.append([p['classifier'], str(p['dataset']), p['_7_point_criteria']['Non-rejected']['precision'], 
            p['_7_point_criteria']['Non-rejected']['specificity'], 
            p['_7_point_criteria']['Non-rejected']['sensitivity'] if isinstance(p['_7_point_criteria']['Non-rejected']['sensitivity'], (int, float)) else 'N/A', 
            p['_7_point_criteria']['Non-rejected']['f1-score'], 
            p['_7_point_criteria']['Non-rejected']['accuracy'],
            p['_7_point_criteria']['Non-rejected']['auc-roc'],
            p['_7_point_criteria']['Threshold'],
            p['_7_point_criteria']['brier_score'],
            p['_7_point_criteria']['ece']])
            
        performance.save(f'{snapshot_path}/performance_reject.xlsx')

        print(f'{snapshot_path}/performance_reject.xlsx' + ' generated')
    
    @staticmethod
    def findOptimalThreshold(jsonfile):

        scores_all = np.array(jsonfile['Val']['y_scores'])
        labels_all = np.array(jsonfile['Val']['y_labels'])

        rej = ClassificationRejector(labels_all, scores_all)
        total_unc = rej.uncertainty(unc_type="TU")

        # Find optimal thresholds
        b_score = []
        eces = []
        avgs = []
        thresholds = np.arange(0.05, 0.2, 0.01)
        for th in thresholds:
            cm = reject_cm(correct=rej.correct, unc_ary=total_unc, threshold=th, relative= True, show=False)

            scores_rejected = scores_all[cm[1] == True]
            labels_rejected = labels_all[cm[1] == True].tolist()
            prob_pos = scores_rejected[:, 1]
            # Get Brier and ECE scores
            brier_score = brier_score_loss(labels_rejected, prob_pos)
            ece_score = mel.Model.expected_calibration_error(scores_all, labels_all, M=10)
            avg = (brier_score + ece_score)/2
            b_score.append(brier_score)
            eces.append(ece_score)
            avgs.append(avg)

        # Find the minimum ECE and its corresponding threshold
        min_avg_index = np.argmin(np.array(avgs))  # Get the index of the minimum ECE score
        optimal_threshold = thresholds[min_avg_index]  # Get the threshold that corresponds to the min ECE score

        return optimal_threshold

        
        

    def reject_uncertainties(snapshot_path):
        jsonfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/ResNet152/performance/*_*_metrics.json', recursive=True)]))
        jsonnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, jsonfiles))
        jsonpaths = list(map(lambda x: pathlib.Path(os.path.dirname(x)), jsonfiles))

        

        plt.ioff()

        DBtypes = [db.name for db in mel.DatasetType]

        for idx, j in enumerate(jsonfiles):
            print(f'Rejecting {j} model')
            fi = open(j)
            jfile = json.load(fi)


            if not os.path.exists(os.path.join(jsonpaths[idx], 'reject_plots')):
                os.makedirs(os.path.join(jsonpaths[idx], 'reject_plots'), exist_ok=True)

            final_uncertainty = {}


            classifier_name = pathlib.Path(jsonpaths[idx]).parent.stem
            used_DB_list = [each_model for each_model in DBtypes if(each_model in jsonnames[idx])]
            DBnames = '+'.join(used_DB_list)
            
            final_uncertainty['dataset'] = used_DB_list
            final_uncertainty['classifier'] = classifier_name
            for phase in ['Val', 'Test']:
                for db in ('ISIC2017', 'ISIC2018', 'KaggleMB', '_7_point_criteria'):
                    print(f"Rejecting {db} database from {jsonnames[idx]}")
                    
                    scores_all = np.array(jfile[db][phase]['y_scores'])
                    labels_all = np.array(jfile[db][phase]['y_labels'])
                    

                    # Instantiate Rejector
                    rej = ClassificationRejector(labels_all, scores_all)
                    # Get entropy (Uncertainty Total (Entropy))
                    # aleatoric - mean of TU (entropy), epistemic: TU - aleatoric
                    total_unc = rej.uncertainty(unc_type="TU")
                    all_unc = rej.uncertainty(unc_type=None)
                    # Plotting uncertainty per test sample
                    uncertainty_plot = rej.plot_uncertainty(unc_type="TU")
                    
                    # uncertainty_plot.suptitle(f'Testset: {db} \nTrainset: {DBnames} \nClassifier: {classifier_name}', fontsize=10, y=1.1, ha='left')
                    if phase == 'Test':
                        preds_all = np.array(jfile[db][phase]['y_pred'])
                        ids_all = np.array(jfile[db][phase]['y_ids'])
                        uncertainty_plot.savefig(os.path.join(jsonpaths[idx], 'reject_plots', DBnames+'_'+classifier_name+'_'+db+'_uncertainty_test'), bbox_inches='tight')
                        # rej.plot_uncertainty(unc_type=None)
                        # implement single rejection point
                        threshold = mel.Model.findOptimalThreshold(jfile[db])
                        
                        reject_output = rej.reject(threshold=threshold, unc_type="TU", relative=True, show=True)

                        cm = reject_cm(correct=rej.correct, unc_ary=total_unc, threshold= threshold,relative= True, show=False)

                        
                        reject_info = {}
                        # False: Not rejected, True: rejected (removed)
                        reject_info['y_labels'] = labels_all.tolist()
                        reject_info['y_preds'] = preds_all.tolist()
                        reject_info['y_scores'] = scores_all.tolist()
                        reject_info['uncertainty'] = total_unc.tolist()
                        reject_info['Threshold'] = threshold
                        
                        reject_info['Non-rejected'] = {}
                        reject_info['Non-rejected']['y_labels'] = labels_all[cm[1] == False].tolist()
                        reject_info['Non-rejected']['y_preds'] = preds_all[cm[1] == False].tolist()
                        reject_info['Non-rejected']['y_scores'] = scores_all[cm[1] == False].tolist()
                        reject_info['Non-rejected']['y_ids'] = ids_all[cm[1] == False].tolist()
                        reject_info['Non-rejected']['Correct'] = {}
                        reject_info['Non-rejected']['Correct']['total'] = cm[0][1]
                        reject_info['Non-rejected']['Correct']['benign(FP)'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 0) & (np.array(reject_info['Non-rejected']['y_preds']) == 0)))
                        reject_info['Non-rejected']['Correct']['malignant(FN)'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 1) & (np.array(reject_info['Non-rejected']['y_preds']) == 1)))
                        reject_info['Non-rejected']['Incorrect'] = {}
                        reject_info['Non-rejected']['Incorrect']['total'] = cm[0][3]
                        reject_info['Non-rejected']['Incorrect']['benign(FP)'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 0) & (np.array(reject_info['Non-rejected']['y_preds']) == 1)))
                        reject_info['Non-rejected']['Incorrect']['malignant(FN)'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 1) & (np.array(reject_info['Non-rejected']['y_preds']) == 0)))

                        reject_info['Rejected'] = {}
                        reject_info['Rejected']['y_labels'] = labels_all[cm[1] == True].tolist()
                        reject_info['Rejected']['y_preds'] = preds_all[cm[1] == True].tolist()
                        reject_info['Rejected']['y_scores'] = scores_all[cm[1] == True].tolist()
                        reject_info['Rejected']['y_ids'] = ids_all[cm[1] == True].tolist()
                        reject_info['Rejected']['Correct'] = {}
                        reject_info['Rejected']['Correct']['total'] = cm[0][0]
                        reject_info['Rejected']['Correct']['benign'] = \
                            int(np.sum((np.array(reject_info['Rejected']['y_labels']) == 0) & (np.array(reject_info['Rejected']['y_preds']) == 0)))
                        reject_info['Rejected']['Correct']['malignant'] = \
                            int(np.sum((np.array(reject_info['Rejected']['y_labels']) == 1) & (np.array(reject_info['Rejected']['y_preds']) == 1)))
                        reject_info['Rejected']['Incorrect'] = {}
                        reject_info['Rejected']['Incorrect']['total'] = cm[0][2]
                        reject_info['Rejected']['Incorrect']['benign(FP)_before'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 0) & (np.array(reject_info['Non-rejected']['y_preds']) == 1))) + \
                                int(np.sum((np.array(reject_info['Rejected']['y_labels']) == 0) & (np.array(reject_info['Rejected']['y_preds']) == 1)))
                        reject_info['Rejected']['Incorrect']['benign(FP)_after'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 0) & (np.array(reject_info['Non-rejected']['y_preds']) == 1)))
                        reject_info['Rejected']['Incorrect']['malignant(FN)_before'] = \
                            int(np.sum((np.array(reject_info['Non-rejected']['y_labels']) == 1) & (np.array(reject_info['Non-rejected']['y_preds']) == 0))) + \
                                int(np.sum((np.array(reject_info['Rejected']['y_labels']) == 1) & (np.array(reject_info['Rejected']['y_preds']) == 0)))
                        reject_info['Rejected']['Incorrect']['malignant(FN)_after'] = \
                            int(np.sum((np.array(reject_info['Rejected']['y_labels']) == 1) & (np.array(reject_info['Rejected']['y_preds']) == 0)))
                        
                        reject_info['Non-rejected_accuracy'] = reject_output[0][0]
                        reject_info['Classification_quality'] = reject_output[0][1]
                        reject_info['Rejection_quality'] = reject_output[0][2]

                        prob_pos = scores_all[cm[1] == False][:, 1]
                    
                        # Brier Score
                        b_score = brier_score_loss(reject_info['Non-rejected']['y_labels'], prob_pos)
                        # print("Brier Score :",b_score)
                        # ECE
                        ece = mel.Model.expected_calibration_error(reject_info['Non-rejected']['y_scores'], reject_info['Non-rejected']['y_labels'], M=10)

                        reject_info['brier_score'] = b_score
                        reject_info['ece'] = ece
                        
                        # Relative threshold
                        threshold_plt = rej.plot_reject(unc_type="TU", metric="NRA")
                        threshold_plt.savefig(os.path.join(jsonpaths[idx], 'reject_plots', DBnames+'_'+classifier_name+'_'+db+'_threshold_test'), bbox_inches='tight')
                        # Absolute threshold
                        # rej.plot_reject(unc_type="TU", metric="NRA", relative=False)

                        common_binary_label = {
                                0.0: 'benign',
                        }

                        mal_prob = [x[1] for x in reject_info['Non-rejected']['y_scores']]
                        if (len(np.unique(reject_info['Non-rejected']['y_labels'])) == 2):
                            test_report = classification_report(reject_info['Non-rejected']['y_labels'], reject_info['Non-rejected']['y_preds'],
                            target_names=mel.Parser.common_binary_label.values(), output_dict = True)

                            reject_info['Non-rejected']['accuracy'] = test_report['accuracy']
                            reject_info['Non-rejected']['precision'] = test_report['macro avg']['precision']
                            reject_info['Non-rejected']['sensitivity'] = test_report['malignant']['recall']
                            reject_info['Non-rejected']['specificity'] = test_report['benign']['recall']
                            reject_info['Non-rejected']['f1-score'] = test_report['macro avg']['f1-score']
                            reject_info['Non-rejected']['auc-roc'] = roc_auc_score(reject_info['Non-rejected']['y_labels'], mal_prob)

                        elif (len(np.unique(reject_info['Non-rejected']['y_labels'])) == 1):
                            test_report = classification_report(reject_info['Non-rejected_y_labels'], reject_info['Non-rejected_y_preds'],
                            target_names=common_binary_label.values(), output_dict = True)

                            reject_info['Non-rejected']['accuracy'] = test_report['accuracy']
                            reject_info['Non-rejected']['precision'] = test_report['macro avg']['precision']
                            reject_info['Non-rejected']['sensitivity'] = '-'
                            reject_info['Non-rejected']['specificity'] = test_report['benign']['recall']
                            reject_info['Non-rejected']['f1-score'] = test_report['macro avg']['f1-score']
                            reject_info['Non-rejected']['auc-roc'] = roc_auc_score(reject_info['Non-rejected']['y_labels'], mal_prob)

                        final_uncertainty[db] = reject_info

                    elif phase == 'Val':
                        # uncertainty_plot.title(db, fontsize=16, color="black", fontweight="bold")
                        if db == '_7_point_criteria':
                            db = '7_point_criteria'
                        elif db == 'KaggleMB':
                            db = 'Kaggle'
                        uncertainty_plot.suptitle(f'{db}', fontsize=15, y=0.95, ha='center')
                        uncertainty_plot.savefig(os.path.join(jsonpaths[idx], 'reject_plots', DBnames+'_'+classifier_name+'_'+db+'_uncertainty_val'), bbox_inches='tight')
                        threshold_plt = rej.plot_reject(unc_type="TU", metric="NRA")
                        threshold_plt.suptitle(f'{db}', fontsize=15, y=0.95, ha='center')
                        # threshold_plt.tick_params(axis='x', labelsize=14)
                        # threshold_plt.supxlabel('Relative threshold', fontsize=14)
                        threshold_plt.savefig(os.path.join(jsonpaths[idx], 'reject_plots', DBnames+'_'+classifier_name+'_'+db+'_threshold_val'), bbox_inches='tight')

                    # final_uncertainty.append(reject_info)

                    


            # Dump Json
            json_filename = f'{jsonnames[idx]}_reject.json'

            file_json = open(os.path.join(jsonpaths[idx], json_filename), "w")
            json.dump(final_uncertainty, file_json, indent = 6)
            file_json.close()

        # y_pred_all, y_true_all = generate_synthetic_output(NUM_SAMPLES, NUM_OBSERVATIONS)
        
        # (num_observations{IN+OOD}, num_samples, NUM_CLASSES)



	
    @staticmethod
    def computing_prediction(model, model_name, target_db, testimages):
        print(f'Computing predictions for {model_name} on {target_db}...')
        # train_pred = model.predict(trainimages)
        # train_pred_classes = np.argmax(train_pred,axis = 1)
        test_pred = model.predict(testimages)
        # Convert predictions classes to one hot vectors
        test_pred_classes = np.argmax(test_pred,axis = 1)

        return test_pred, test_pred_classes

    @staticmethod
    def evaluate_model_json(all_labels, all_preds, all_scores, snapshot_path):
        test_report = classification_report(all_labels, all_preds, target_names=mel.Parser.common_binary_label.values(), output_dict = True)
        

        return test_report
    
    @staticmethod
    def ensemble(snapshot_path):
        jsonfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/*/performance/*_metrics.json', recursive=True)]))
        jsonnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, jsonfiles))

        performances = []


        
        test_sets = ['ISIC2017', 'ISIC2018', '_7_point_criteria', 'KaggleMB']
        phases = ['Test']
        
        ensemble = {}
        for db in test_sets:
            
            ensemble.setdefault(db, {})
            for phase in phases: # for phase in ['Val', 'Test']:
                ensemble[db].setdefault(phase, {})
                for idx, j in enumerate(jsonfiles):
                    fi = open(j)
                    jfile = json.load(fi)
                    ensemble[db][phase].setdefault('y_labels', jfile[db][phase]['y_labels'])
                    ensemble[db][phase].setdefault('combinations', {})
                    ensemble[db][phase]['combinations'].setdefault(idx, {})
                    ensemble[db][phase]['combinations'][idx].setdefault('combination', jsonnames[idx])
                    ensemble[db][phase]['combinations'][idx].setdefault('y_pred', jfile[db][phase]['y_pred'])
                    ensemble[db][phase]['combinations'][idx].setdefault('y_scores', jfile[db][phase]['y_scores'])
                    ensemble[db][phase]['combinations'][idx].setdefault('classifier', jfile['classifier'])
                # Initialization
                ensemble[db][phase].setdefault('counts', [0] * len(ensemble[db][phase]['combinations'][0]['y_pred']))
                ensemble[db][phase].setdefault('probs', [0] * len(ensemble[db][phase]['combinations'][0]['y_scores']))
                ensemble[db][phase].setdefault('probs_weighted', [0] * len(ensemble[db][phase]['combinations'][0]['y_scores']))

        
        statistics = mel.Model.extract_positives_per_sample(snapshot_path)
        total_models = sum(statistics.values())
        model_weights = {model_name: count / total_models for model_name, count in statistics.items()}
        test_report = {}
        performance = {}
        for db in test_sets:
            for phase in phases:
                for comb in ensemble[db][phase]['combinations']:
                    y_label = ensemble[db][phase]['y_labels']
                    y_pred = ensemble[db][phase]['combinations'][comb]['y_pred']
                    y_scores = ensemble[db][phase]['combinations'][comb]['y_scores']
                    classifier = ensemble[db][phase]['combinations'][comb]['classifier'] # network name
                    
                    ensemble[db][phase]['counts'] = [
                        count + (1 if pred == 1 else 0)
                        for count, pred in zip(
                            ensemble[db][phase]['counts'], y_pred
                        )
                    ]
                    # ensemble[db]['probs'] = [score + pred for score, pred in zip(ensemble[db]['probs'], y_scores[db][1])]
                    ensemble[db][phase]['probs'] = [
                        score + prob[1]/len(ensemble[db][phase]['combinations'])
                        for score, prob in zip(ensemble[db][phase]['probs'], y_scores)
                    ]

                    ensemble[db][phase]['probs_weighted'] = [
                        score + (prob[1]/len(ensemble[db][phase]['combinations']))*model_weights[classifier]
                        for score, prob in zip(ensemble[db][phase]['probs_weighted'], y_scores)
                    ]

                    # ensemble[db][phase]['probs_weighted']= [
                    #     score + (model_weights[model_name] * prob[1]) 
                    #     for score, prob, model_name in zip(
                    #         ensemble[db][phase]['probs'], 
                    #         y_scores, 
                    #         ensemble[db][phase]['combinations']['classifier']
                    #     )
                    # ]
                    # for idx, prob in enumerate(y_scores):
                    #     classifier_name = ensemble[db][phase]['combinations'][idx]['classifier']
                    #     weight = model_weights[classifier_name]  # Get the weight for the current classifier
                    #     ensemble[db][phase]['probs_weighted'] = [
                    #         score + (weight * p)
                    #         for score, p in zip(ensemble[db][phase]['probs'], prob[1])  # Apply weights to prob[1]
                        # ]
                    
            
                ensemble[db][phase].setdefault('hard_voting', {})
                ensemble[db][phase].setdefault('soft_voting', {})
                for idx, (count, prob) in enumerate(zip(ensemble[db][phase]['counts'], ensemble[db][phase]['probs'])):
                    if count > len(ensemble[db][phase]['counts']) / 2:  # Hard-voting condition
                        ensemble[db][phase]['hard_voting'][idx] = 1
                    else:
                        ensemble[db][phase]['hard_voting'][idx] = 0

                    if prob > 0.5:  # Soft-voting condition
                        ensemble[db][phase]['soft_voting'][idx] = 1
                    else:
                        ensemble[db][phase]['soft_voting'][idx] = 0
                
                
                test_report.setdefault(db, {})
                test_report[db].setdefault(phase, {})
                
                mal_prob = [x[1] for x in y_scores]
                print(f"(Hard voting)Testing Ensemble on {db}")
                test_report[db][phase]['hard_voting'] = classification_report(y_label, list(ensemble[db][phase]['hard_voting'].values()), target_names=mel.Parser.common_binary_label.values(), output_dict = True)
                # test_report[phase]['hard_voting'] = mel.Model.evaluate_model_json(y_label, y_pred, list(ensemble[db][phase]['hard_voting'].values()))
                print(f"(Soft voting)Testing Ensemble on {db}")
                # test_report[phase]['soft_voting'] = mel.Model.evaluate_model_json(y_label, y_pred, list(ensemble[db][phase]['soft_voting'].values()))
                test_report[db][phase]['soft_voting'] = classification_report(y_label, list(ensemble[db][phase]['soft_voting'].values()), target_names=mel.Parser.common_binary_label.values(), output_dict = True)

                performance.setdefault(db, {})
                performance[db].setdefault(phase, {})
                performance[db][phase].setdefault('hard_voting', {})
                performance[db][phase].setdefault('soft_voting', {})
                performance[db][phase]['hard_voting'] = {
                    
                    
                    'y_labels': y_label,
                    'y_pred': list(ensemble[db][phase]['hard_voting'].values()),
                    # 'y_scores': y_scores,
                    # 'y_ids': all_ids['Test'],
                    'accuracy': test_report[db][phase]['hard_voting']['accuracy'],
                    'precision': test_report[db][phase]['hard_voting']['macro avg']['precision'],
                    'sensitivity': test_report[db][phase]['hard_voting']['malignant']['recall'],
                    'specificity': test_report[db][phase]['hard_voting']['benign']['recall'],
                    'f1-score': test_report[db][phase]['hard_voting']['macro avg']['f1-score'],
                    # 'auc-roc': roc_auc_score(all_labels['Test'], mal_prob),
                    # 'brier_score': b_score_test,
                    # 'ece': ece_test,
                    
                }

                performance[db][phase]['soft_voting'] = {
                    
                    
                    'y_labels': y_label,
                    'y_pred': list(ensemble[db][phase]['soft_voting'].values()),
                    # 'y_scores': y_scores,
                    # 'y_ids': all_ids['Test'],
                    'accuracy': test_report[db][phase]['soft_voting']['accuracy'],
                    'precision': test_report[db][phase]['soft_voting']['macro avg']['precision'],
                    'sensitivity': test_report[db][phase]['soft_voting']['malignant']['recall'],
                    'specificity': test_report[db][phase]['soft_voting']['benign']['recall'],
                    'f1-score': test_report[db][phase]['soft_voting']['macro avg']['f1-score'],
                    # 'auc-roc': roc_auc_score(all_labels['Test'], mal_prob),
                    # 'brier_score': b_score_test,
                    # 'ece': ece_test,
                    
                }
        
        print()


    @staticmethod
    def model_report(
        model_path,
        model_name,
        target_db,
        target_network,
        testlabels,
        test_pred_classes,
    ):
        label_substitution = {
			0.0: 'Benign',
			1.0: 'Malignant'
		}
        
        # trainlabels_digit = np.argmax(trainlabels, axis=1)
        testlabels_digit = np.argmax(testlabels, axis=1)

        # train_report = classification_report(trainlabels_digit, train_pred_classes, target_names=label_substitution.values(), output_dict = True)
        test_report = classification_report(testlabels_digit, test_pred_classes, target_names=label_substitution.values(), output_dict = True)

        print(f'Model report for {model_name} model ->\n\n')
        # print("Train Report :\n", train_report)
        print("Test Report :\n", test_report)

        cm = confusion_matrix(testlabels_digit, test_pred_classes)
        
        plt.ioff()
        fig = plt.figure(figsize=(12, 8))
        df_cm = pd.DataFrame(cm, index=label_substitution.values(), columns=label_substitution.values())

        try:
            heatmap = sns.heatmap(df_cm, annot=True, fmt="d", cbar=False, cmap='Blues')
        except ValueError:
            raise ValueError("Confusion matrix values must be integers.")

        heatmap.yaxis.set_ticklabels(heatmap.yaxis.get_ticklabels(), rotation=0, ha='right', fontsize=13)
        heatmap.xaxis.set_ticklabels(heatmap.xaxis.get_ticklabels(), rotation=45, ha='right', fontsize=13)
        plt.ylabel('True label', fontsize=13)
        plt.xlabel('Predicted label', fontsize=13)

        plt.tight_layout()
        plt.subplots_adjust(top=0.95)
        plt.title(f'Confusion Matrix of ({model_name}) on {target_db}', fontsize=13)
        # plt.show()
        if not os.path.exists(f'{model_path}/performance/{target_network}'):
            os.makedirs(f'{model_path}/performance/{target_network}', exist_ok=True)
        plt.savefig(f'{model_path}/performance/{target_network}/{model_name}_confusion1.png')
        conf_mat_df = pd.crosstab(testlabels_digit, test_pred_classes, rownames=['GT Label'],colnames=['Predict'])
        df_styled = conf_mat_df.style.background_gradient()
        dfi.export(df_styled, f"{model_path}/performance/{target_network}/{model_name}_confusion2.png", table_conversion="matplotlib")
        # conf_mat_df.dfi.export(f"{model_path}/performance/{model_name}_confusion2.png")

        plt.close()

        return test_report

    @staticmethod
    def expected_calibration_error(samples, true_labels, M=5):
        # uniform binning approach with M number of bins
        bin_boundaries = np.linspace(0, 1, M + 1)
        bin_lowers = bin_boundaries[:-1]
        bin_uppers = bin_boundaries[1:]

        # get max probability per sample i
        confidences = np.max(samples, axis=1)
        # get predictions from confidences (positional in this case)
        predicted_label = np.argmax(samples, axis=1)

        # get a boolean list of correct/false predictions
        accuracies = predicted_label==true_labels

        ece = np.zeros(1)
        for bin_lower, bin_upper in zip(bin_lowers, bin_uppers):
            # determine if sample is in bin m (between bin lower & upper)
            in_bin = np.logical_and(confidences > bin_lower.item(), confidences <= bin_upper.item())
            # can calculate the empirical probability of a sample falling into bin m: (|Bm|/n)
            prob_in_bin = in_bin.mean()

            if prob_in_bin.item() > 0:
                # get the accuracy of bin m: acc(Bm)
                accuracy_in_bin = accuracies[in_bin].mean()
                # get the average confidence of bin m: conf(Bm)
                avg_confidence_in_bin = confidences[in_bin].mean()
                # calculate |acc(Bm) - conf(Bm)| * (|Bm|/n) for bin m and add to the total ECE
                ece += np.abs(avg_confidence_in_bin - accuracy_in_bin) * prob_in_bin
        return ece[0]
