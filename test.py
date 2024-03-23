import melanoma as mel

import itertools
import glob
import os
os.environ["HDF5_USE_FILE_LOCKING"] = "FALSE"
import pathlib
import json
import openpyxl

CFG={
    'verbose': 1,
}

base_model = mel.CNN(CFG=CFG)

# snapshot_path = '/raid/mpsych/MELANOMA/snapshot/meshnet'
snapshot_path = '/hpcstor6/scratch01/s/sanghyuk.kim001/snapshot/bestperformers'

modelfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/*.hdf5', recursive=True)]))
modelnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, modelfiles))
# modelpath = os.path.abspath(os.path.join(modelfiles, os.pardir))

img_size = (150, 150)
commondata = mel.CommonData()

networktypes = [net.name for net in mel.NetworkType]
modelnames.reverse()

final_perf = []

for idx, m in enumerate(modelnames):
    netname = [each_m for each_m in networktypes if(each_m in m)]
    # assert len(netname) == 1
    if len(netname) == 1:
        mapped_netname = commondata.DBpreprocessorDict[netname[0]]
    elif len(netname) == 2:
        mapped_netname = commondata.DBpreprocessorDict[netname[1]]
    else:
        assert len(netname) <= 2
    dbpath = f'/hpcstor6/scratch01/s/sanghyuk.kim001/melanomaDB/customDB/{mapped_netname}'

    dbpath_KaggleDB = dbpath+'/'+f'KaggleMB_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    dbpath_HAM10000 = dbpath+'/'+f'HAM10000_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    dbpath_ISIC2016 = dbpath+'/'+f'ISIC2016_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    dbpath_ISIC2017 = dbpath+'/'+f'ISIC2017_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    dbpath_ISIC2018 = dbpath+'/'+f'ISIC2018_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    dbpath_ISIC2020 = dbpath+'/'+f'ISIC2020_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    dbpath_7pointcriteria = dbpath+'/'+f'_7_point_criteria_{img_size[0]}h_{img_size[1]}w_binary.pkl'
    # final_perf.append(base_model.evaluate_model_onAll(model_name=m, model_path=snapshot_path, network_name = netname[0], dbpath_KaggleDB=dbpath_KaggleDB, dbpath_HAM10000=dbpath_HAM10000, \
    #     dbpath_ISIC2016=dbpath_ISIC2016, dbpath_ISIC2017=dbpath_ISIC2017, dbpath_ISIC2018=dbpath_ISIC2018, dbpath_7pointcriteria=dbpath_7pointcriteria))
    from tensorflow.keras.models import load_model
    model = load_model(snapshot_path+'/'+m + '.hdf5')
    target_network = model.layers[0].name

    base_model.evaluate_model_onAll(model_name=m, model_path=snapshot_path, network_name = netname[0], dbpath_KaggleDB=dbpath_KaggleDB, dbpath_HAM10000=dbpath_HAM10000, \
        dbpath_ISIC2016=dbpath_ISIC2016, dbpath_ISIC2017=dbpath_ISIC2017, dbpath_ISIC2018=dbpath_ISIC2018, dbpath_7pointcriteria=dbpath_7pointcriteria, dbpath_ISIC2020=dbpath_ISIC2020,
        leaderboard_only = True)




# jsonfiles = list(itertools.chain.from_iterable([glob.glob(f'{snapshot_path}/performance/conv2d/*.json', recursive=True)]))
# jsonnames = list(map(lambda x: pathlib.Path(os.path.basename(x)).stem, jsonfiles))

# for idx, j in enumerate(jsonfiles):
#     fi = open(j)
#     jfile = json.load(fi)
#     final_perf.append(jfile)


# # KaggleMB

# KaggleMB_maxperf = {
#     "max_accuracy": max(final_perf, key=(lambda item: item['KaggleMB']['accuracy'])),
#     "max_precision": max(final_perf, key=(lambda item: item['KaggleMB']['precision'])),
#     "max_sensitivity": max(final_perf, key=(lambda item: item['KaggleMB']['sensitivity'])),
#     "max_specificity": max(final_perf, key=(lambda item: item['KaggleMB']['specificity'])),
#     "max_f1-score": max(final_perf, key=(lambda item: item['KaggleMB']['f1-score'])),
# }

# KaggleMB_maxacc = {
#     "datasets": KaggleMB_maxperf['max_accuracy']['dataset'],
#     "classifier": KaggleMB_maxperf['max_accuracy']['classifier'],
#     "acc": KaggleMB_maxperf['max_accuracy']['KaggleMB']['accuracy'],
# }

# KaggleMB_maxprec = {
#     "datasets": KaggleMB_maxperf['max_precision']['dataset'],
#     "classifier": KaggleMB_maxperf['max_precision']['classifier'],
#     "acc": KaggleMB_maxperf['max_precision']['KaggleMB']['precision'],
# }

# KaggleMB_maxsens = {
#     "datasets": KaggleMB_maxperf['max_sensitivity']['dataset'],
#     "classifier": KaggleMB_maxperf['max_sensitivity']['classifier'],
#     "acc": KaggleMB_maxperf['max_sensitivity']['KaggleMB']['sensitivity'],
# }

# KaggleMB_maxspec = {
#     "datasets": KaggleMB_maxperf['max_specificity']['dataset'],
#     "classifier": KaggleMB_maxperf['max_specificity']['classifier'],
#     "acc": KaggleMB_maxperf['max_specificity']['KaggleMB']['specificity'],
# }

# KaggleMB_maxf1 = {
#     "datasets": KaggleMB_maxperf['max_accuracy']['dataset'],
#     "classifier": KaggleMB_maxperf['max_accuracy']['classifier'],
#     "acc": KaggleMB_maxperf['max_f1-score']['KaggleMB']['f1-score'],
# }


# # HAM10000

# HAM10000_maxperf = {
#     "max_accuracy": max(final_perf, key=(lambda item: item['HAM10000']['accuracy'])),
#     "max_precision": max(final_perf, key=(lambda item: item['HAM10000']['precision'])),
#     "max_sensitivity": max(final_perf, key=(lambda item: item['HAM10000']['sensitivity'])),
#     "max_specificity": max(final_perf, key=(lambda item: item['HAM10000']['specificity'])),
#     "max_f1-score": max(final_perf, key=(lambda item: item['HAM10000']['f1-score'])),
# }

# HAM10000_maxacc = {
#     "datasets": HAM10000_maxperf['max_accuracy']['dataset'],
#     "classifier": HAM10000_maxperf['max_accuracy']['classifier'],
#     "acc": HAM10000_maxperf['max_accuracy']['HAM10000']['accuracy'],
# }

# HAM10000_maxprec = {
#     "datasets": HAM10000_maxperf['max_precision']['dataset'],
#     "classifier": HAM10000_maxperf['max_precision']['classifier'],
#     "acc": HAM10000_maxperf['max_precision']['HAM10000']['precision'],
# }

# HAM10000_maxsens = {
#     "datasets": HAM10000_maxperf['max_sensitivity']['dataset'],
#     "classifier": HAM10000_maxperf['max_sensitivity']['classifier'],
#     "acc": HAM10000_maxperf['max_sensitivity']['HAM10000']['sensitivity'],
# }

# HAM10000_maxspec = {
#     "datasets": HAM10000_maxperf['max_specificity']['dataset'],
#     "classifier": HAM10000_maxperf['max_specificity']['classifier'],
#     "acc": HAM10000_maxperf['max_specificity']['HAM10000']['specificity'],
# }

# HAM10000_maxf1 = {
#     "datasets": HAM10000_maxperf['max_accuracy']['dataset'],
#     "classifier": HAM10000_maxperf['max_accuracy']['classifier'],
#     "acc": HAM10000_maxperf['max_f1-score']['HAM10000']['f1-score'],
# }

# # ISIC2016

# ISIC2016_maxperf = {
#     "max_accuracy": max(final_perf, key=(lambda item: item['ISIC2016']['accuracy'])),
#     "max_precision": max(final_perf, key=(lambda item: item['ISIC2016']['precision'])),
#     "max_sensitivity": max(final_perf, key=(lambda item: item['ISIC2016']['sensitivity'])),
#     "max_specificity": max(final_perf, key=(lambda item: item['ISIC2016']['specificity'])),
#     "max_f1-score": max(final_perf, key=(lambda item: item['ISIC2016']['f1-score'])),
# }

# ISIC2016_maxacc = {
#     "datasets": ISIC2016_maxperf['max_accuracy']['dataset'],
#     "classifier": ISIC2016_maxperf['max_accuracy']['classifier'],
#     "acc": ISIC2016_maxperf['max_accuracy']['ISIC2016']['accuracy'],
# }

# ISIC2016_maxprec = {
#     "datasets": ISIC2016_maxperf['max_precision']['dataset'],
#     "classifier": ISIC2016_maxperf['max_precision']['classifier'],
#     "acc": ISIC2016_maxperf['max_precision']['ISIC2016']['precision'],
# }

# ISIC2016_maxsens = {
#     "datasets": ISIC2016_maxperf['max_sensitivity']['dataset'],
#     "classifier": ISIC2016_maxperf['max_sensitivity']['classifier'],
#     "acc": ISIC2016_maxperf['max_sensitivity']['ISIC2016']['sensitivity'],
# }

# ISIC2016_maxspec = {
#     "datasets": ISIC2016_maxperf['max_specificity']['dataset'],
#     "classifier": ISIC2016_maxperf['max_specificity']['classifier'],
#     "acc": ISIC2016_maxperf['max_specificity']['ISIC2016']['specificity'],
# }

# ISIC2016_maxf1 = {
#     "datasets": ISIC2016_maxperf['max_accuracy']['dataset'],
#     "classifier": ISIC2016_maxperf['max_accuracy']['classifier'],
#     "acc": ISIC2016_maxperf['max_f1-score']['ISIC2016']['f1-score'],
# }

# # ISIC2017

# ISIC2017_maxperf = {
#     "max_accuracy": max(final_perf, key=(lambda item: item['ISIC2017']['accuracy'])),
#     "max_precision": max(final_perf, key=(lambda item: item['ISIC2017']['precision'])),
#     "max_sensitivity": max(final_perf, key=(lambda item: item['ISIC2017']['sensitivity'])),
#     "max_specificity": max(final_perf, key=(lambda item: item['ISIC2017']['specificity'])),
#     "max_f1-score": max(final_perf, key=(lambda item: item['ISIC2017']['f1-score'])),
# }

# ISIC2017_maxacc = {
#     "datasets": ISIC2017_maxperf['max_accuracy']['dataset'],
#     "classifier": ISIC2017_maxperf['max_accuracy']['classifier'],
#     "acc": ISIC2017_maxperf['max_accuracy']['ISIC2017']['accuracy'],
# }

# ISIC2017_maxprec = {
#     "datasets": ISIC2017_maxperf['max_precision']['dataset'],
#     "classifier": ISIC2017_maxperf['max_precision']['classifier'],
#     "acc": ISIC2017_maxperf['max_precision']['ISIC2017']['precision'],
# }

# ISIC2017_maxsens = {
#     "datasets": ISIC2017_maxperf['max_sensitivity']['dataset'],
#     "classifier": ISIC2017_maxperf['max_sensitivity']['classifier'],
#     "acc": ISIC2017_maxperf['max_sensitivity']['ISIC2017']['sensitivity'],
# }

# ISIC2017_maxspec = {
#     "datasets": ISIC2017_maxperf['max_specificity']['dataset'],
#     "classifier": ISIC2017_maxperf['max_specificity']['classifier'],
#     "acc": ISIC2017_maxperf['max_specificity']['ISIC2017']['specificity'],
# }

# ISIC2017_maxf1 = {
#     "datasets": ISIC2017_maxperf['max_accuracy']['dataset'],
#     "classifier": ISIC2017_maxperf['max_accuracy']['classifier'],
#     "acc": ISIC2017_maxperf['max_f1-score']['ISIC2017']['f1-score'],
# }

# # ISIC2018

# ISIC2018_maxperf = {
#     "max_accuracy": max(final_perf, key=(lambda item: item['ISIC2018']['accuracy'])),
#     "max_precision": max(final_perf, key=(lambda item: item['ISIC2018']['precision'])),
#     "max_sensitivity": max(final_perf, key=(lambda item: item['ISIC2018']['sensitivity'])),
#     "max_specificity": max(final_perf, key=(lambda item: item['ISIC2018']['specificity'])),
#     "max_f1-score": max(final_perf, key=(lambda item: item['ISIC2018']['f1-score'])),
# }

# ISIC2018_maxacc = {
#     "datasets": ISIC2018_maxperf['max_accuracy']['dataset'],
#     "classifier": ISIC2018_maxperf['max_accuracy']['classifier'],
#     "acc": ISIC2018_maxperf['max_accuracy']['ISIC2018']['accuracy'],
# }

# ISIC2018_maxprec = {
#     "datasets": ISIC2018_maxperf['max_precision']['dataset'],
#     "classifier": ISIC2018_maxperf['max_precision']['classifier'],
#     "acc": ISIC2018_maxperf['max_precision']['ISIC2018']['precision'],
# }

# ISIC2018_maxsens = {
#     "datasets": ISIC2018_maxperf['max_sensitivity']['dataset'],
#     "classifier": ISIC2018_maxperf['max_sensitivity']['classifier'],
#     "acc": ISIC2018_maxperf['max_sensitivity']['ISIC2018']['sensitivity'],
# }

# ISIC2018_maxspec = {
#     "datasets": ISIC2018_maxperf['max_specificity']['dataset'],
#     "classifier": ISIC2018_maxperf['max_specificity']['classifier'],
#     "acc": ISIC2018_maxperf['max_specificity']['ISIC2018']['specificity'],
# }

# ISIC2018_maxf1 = {
#     "datasets": ISIC2018_maxperf['max_accuracy']['dataset'],
#     "classifier": ISIC2018_maxperf['max_accuracy']['classifier'],
#     "acc": ISIC2018_maxperf['max_f1-score']['ISIC2018']['f1-score'],
# }


# performance = openpyxl.Workbook()
# performance_ws = performance.active
# performance_ws.title = 'HAM10000'
# performance.create_sheet('ISIC2016')
# performance.create_sheet('ISIC2017')
# performance.create_sheet('ISIC2018')
# performance.create_sheet('KaggleMB')
# performance.create_sheet('7pointcriteria')


# cols = ['net/metrics', 'DB Comb', 'Precision', 'Specificity', 'Sensitivity', 'Accuracy', 'Filesize', 'Parameters']

# HAM10000_ws = performance['HAM10000']
# HAM10000_ws.append(cols)

# for p in final_perf:
#     HAM10000_ws.append([p['classifier'], str(p['dataset']), p['HAM10000']['precision'], p['HAM10000']['specificity'], p['HAM10000']['sensitivity'], p['HAM10000']['accuracy'], p['Filesize'], p['Parameters']])
    
# ISIC2016_ws = performance['ISIC2016']
# ISIC2016_ws.append(cols)

# for p in final_perf:
#     ISIC2016_ws.append([p['classifier'], str(p['dataset']), p['ISIC2016']['precision'], p['ISIC2016']['specificity'], p['ISIC2016']['sensitivity'], p['ISIC2016']['accuracy'], p['Filesize'], p['Parameters']])

# ISIC2017_ws = performance['ISIC2017']
# ISIC2017_ws.append(cols)

# for p in final_perf:
#     ISIC2017_ws.append([p['classifier'], str(p['dataset']), p['ISIC2017']['precision'], p['ISIC2017']['specificity'], p['ISIC2017']['sensitivity'], p['ISIC2017']['accuracy'], p['Filesize'], p['Parameters']])

# ISIC2018_ws = performance['ISIC2018']
# ISIC2018_ws.append(cols)

# for p in final_perf:
#     ISIC2018_ws.append([p['classifier'], str(p['dataset']), p['ISIC2018']['precision'], p['ISIC2018']['specificity'], p['ISIC2018']['sensitivity'], p['ISIC2018']['accuracy'], p['Filesize'], p['Parameters']])


# KaggleMB_ws = performance['KaggleMB']
# KaggleMB_ws.append(cols)

# for p in final_perf:
#     KaggleMB_ws.append([p['classifier'], str(p['dataset']), p['KaggleMB']['precision'], p['KaggleMB']['specificity'], p['KaggleMB']['sensitivity'], p['KaggleMB']['accuracy'], p['Filesize'], p['Parameters']])

# _7pointcriteria_ws = performance['7pointcriteria']
# _7pointcriteria_ws.append(cols)

# for p in final_perf:
#     _7pointcriteria_ws.append([p['classifier'], str(p['dataset']), p['_7_point_criteria']['precision'], p['_7_point_criteria']['specificity'], p['_7_point_criteria']['sensitivity'], p['_7_point_criteria']['accuracy'], p['Filesize'], p['Parameters']])



    
# performance.save(f'{snapshot_path}/performance/performance.xlsx')


print("finish")